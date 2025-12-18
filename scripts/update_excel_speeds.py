"""
Excel dosyasindaki hiz degerlerini S3'deki guncel summary.csv'lerden gunceller.
"""
import sys
import re
import boto3
import pandas as pd
from openpyxl import load_workbook
from io import StringIO

sys.stdout.reconfigure(encoding='utf-8')

S3_BUCKET = 'microplastic-experiments'
s3 = boto3.client('s3')

EXCEL_PATH = 'c:/Users/mmert/PycharmProjects/ObjectTrackingProject/Video_Boyut_Eslestirme_FINAL.xlsx'
OUTPUT_PATH = 'c:/Users/mmert/PycharmProjects/ObjectTrackingProject/Video_Boyut_Eslestirme_FINAL_updated.xlsx'

def get_speed_from_s3(view, date, repeat, category, code):
    """S3'den hiz degerini al"""
    # Tarih formatini duzelt
    date_formatted = date.replace('.', '.')

    # S3 path olustur - once success, sonra fail dene
    for status in ['success', 'fail']:
        for fail_reason in ['', 'insufficient_movement/', 'not_found/', 'lost_early/', 'video_error/', 'horizontal_drift/']:
            if status == 'success' and fail_reason:
                continue

            prefix = f"{status}/{fail_reason}{date_formatted}/{view}/{repeat}/{category}/{code}/"

            try:
                response = s3.list_objects_v2(Bucket=S3_BUCKET, Prefix=prefix, MaxKeys=1)
                if response.get('Contents'):
                    summary_key = f"{prefix}summary.csv"
                    try:
                        obj = s3.get_object(Bucket=S3_BUCKET, Key=summary_key)
                        content = obj['Body'].read().decode('utf-8')

                        for line in content.split('\n'):
                            if line.startswith('Hiz,'):
                                parts = line.split(',')
                                if len(parts) >= 2:
                                    speed = float(parts[1])
                                    # cm/s ise m/s'e cevir
                                    if speed > 1:
                                        speed = speed / 100
                                    return speed
                    except:
                        pass
            except:
                pass

    return None

def parse_klasor(klasor_str):
    """Klasor adından tarih ve view bilgisini çıkar"""
    # IC CAPTURE 01.11.23(MAK) -> 01.11.23, MAK
    # IC CAPTURE 01.11.23-ANG -> 01.11.23, ANG
    # IC CAPTURE 4.12.2024-ANG -> 04.12.2024, ANG

    klasor_str = str(klasor_str)

    # View'u bul
    view = None
    if '(MAK)' in klasor_str or '-MAK' in klasor_str:
        view = 'MAK'
    elif '(ANG)' in klasor_str or '-ANG' in klasor_str:
        view = 'ANG'
    elif 'MAK' in klasor_str.upper():
        view = 'MAK'
    elif 'ANG' in klasor_str.upper():
        view = 'ANG'

    # Tarihi bul
    date_match = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})', klasor_str)
    if date_match:
        day, month, year = date_match.groups()
        day = day.zfill(2)
        month = month.zfill(2)
        if len(year) == 2:
            year = '20' + year if int(year) < 50 else '19' + year
        date = f"{day}.{month}.{year}"
    else:
        date = None

    return date, view

def update_excel():
    print("Excel dosyasi yukleniyor...")
    wb = load_workbook(EXCEL_PATH)

    total_updated = 0
    total_not_found = 0

    for sheet_name in ['MAK', 'ANG']:
        print(f"\n=== {sheet_name} sayfasi isleniyor ===")
        ws = wb[sheet_name]

        # Kolon indekslerini bul
        headers = [cell.value for cell in ws[1]]
        hiz_col = headers.index('Hiz (m/s)') + 1 if 'Hiz (m/s)' in headers else None
        klasor_col = headers.index('Klasor') + 1 if 'Klasor' in headers else None
        deney_col = headers.index('Deney') + 1 if 'Deney' in headers else None
        kategori_col = headers.index('Kategori') + 1 if 'Kategori' in headers else None
        kod_col = headers.index('Kod') + 1 if 'Kod' in headers else None

        if not all([hiz_col, klasor_col, kategori_col, kod_col]):
            print(f"  Gerekli kolonlar bulunamadi: {headers}")
            continue

        updated = 0
        not_found = 0

        for row in range(2, ws.max_row + 1):
            klasor = ws.cell(row=row, column=klasor_col).value
            deney = ws.cell(row=row, column=deney_col).value if deney_col else None
            kategori = ws.cell(row=row, column=kategori_col).value
            kod = ws.cell(row=row, column=kod_col).value
            old_speed = ws.cell(row=row, column=hiz_col).value

            if not klasor or not kategori or not kod:
                continue

            date, view = parse_klasor(klasor)
            if not date or not view:
                continue

            # Deney (FIRST, SECOND, THIRD)
            repeat = str(deney).upper() if deney else 'FIRST'

            # S3'den yeni hizi al
            new_speed = get_speed_from_s3(view, date, repeat, kategori, kod)

            if new_speed is not None:
                ws.cell(row=row, column=hiz_col).value = round(new_speed, 4)
                updated += 1
                if updated <= 5:
                    print(f"  {date}/{view}/{repeat}/{kategori}/{kod}: {old_speed} -> {new_speed:.4f}")
            else:
                not_found += 1

            if (row - 1) % 100 == 0:
                print(f"  Ilerleme: {row-1}/{ws.max_row-1}")

        print(f"  Guncellenen: {updated}, Bulunamayan: {not_found}")
        total_updated += updated
        total_not_found += not_found

    # Ortalama sayfalarini da guncelle
    for sheet_name in ['ANG Ortalama', 'MAK Ortalama']:
        print(f"\n=== {sheet_name} sayfasi isleniyor ===")
        ws = wb[sheet_name]

        headers = [cell.value for cell in ws[1]]
        hiz_col = headers.index('Hiz (m/s)') + 1 if 'Hiz (m/s)' in headers else None
        klasor_col = headers.index('Klasor') + 1 if 'Klasor' in headers else None
        kategori_col = headers.index('Kategori') + 1 if 'Kategori' in headers else None
        kod_col = headers.index('Kod') + 1 if 'Kod' in headers else None

        if not all([hiz_col, klasor_col, kategori_col, kod_col]):
            print(f"  Gerekli kolonlar bulunamadi")
            continue

        updated = 0
        not_found = 0

        for row in range(2, ws.max_row + 1):
            klasor = ws.cell(row=row, column=klasor_col).value
            kategori = ws.cell(row=row, column=kategori_col).value
            kod = ws.cell(row=row, column=kod_col).value
            old_speed = ws.cell(row=row, column=hiz_col).value

            if not klasor or not kategori or not kod:
                continue

            date, view = parse_klasor(klasor)
            if not date or not view:
                continue

            # Ortalama sayfalarda 3 tekrarin ortalamasini al
            speeds = []
            for repeat in ['FIRST', 'SECOND', 'THIRD']:
                speed = get_speed_from_s3(view, date, repeat, kategori, kod)
                if speed is not None:
                    speeds.append(speed)

            if speeds:
                avg_speed = sum(speeds) / len(speeds)
                ws.cell(row=row, column=hiz_col).value = round(avg_speed, 4)
                updated += 1
                if updated <= 3:
                    print(f"  {date}/{view}/{kategori}/{kod}: {old_speed} -> {avg_speed:.4f} (n={len(speeds)})")
            else:
                not_found += 1

            if (row - 1) % 50 == 0:
                print(f"  Ilerleme: {row-1}/{ws.max_row-1}")

        print(f"  Guncellenen: {updated}, Bulunamayan: {not_found}")
        total_updated += updated
        total_not_found += not_found

    print(f"\n=== TOPLAM ===")
    print(f"Guncellenen: {total_updated}")
    print(f"Bulunamayan: {total_not_found}")

    print(f"\nKaydediliyor: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    print("Tamamlandi!")

if __name__ == '__main__':
    update_excel()
