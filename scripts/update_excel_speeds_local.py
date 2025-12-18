"""
Excel dosyasindaki hiz degerlerini LOCAL summary.csv'lerden gunceller.
"""
import sys
import os
import re
import glob
import pandas as pd
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = 'c:/Users/mmert/PycharmProjects/ObjectTrackingProject/Video_Boyut_Eslestirme_FINAL.xlsx'
OUTPUT_PATH = 'c:/Users/mmert/PycharmProjects/ObjectTrackingProject/Video_Boyut_Eslestirme_FINAL_updated.xlsx'
RESULTS_DIR = 'c:/Users/mmert/PycharmProjects/ObjectTrackingProject/processed_results'

# Tum summary.csv'leri bir kere oku ve cache'le
speed_cache = {}

def build_speed_cache():
    """Tum summary.csv dosyalarini oku ve cache'e al"""
    print("Summary dosyalari okunuyor...")

    # glob ile tum summary.csv dosyalarini bul
    pattern = os.path.join(RESULTS_DIR, '**', 'summary.csv')
    summary_files = glob.glob(pattern, recursive=True)
    print(f"Bulunan summary dosyasi: {len(summary_files)}")

    for summary_path in summary_files:
        # Path'den bilgileri cikar
        rel_path = os.path.relpath(summary_path, RESULTS_DIR)
        parts = rel_path.replace('\\', '/').split('/')

        # success/DATE/VIEW/REPEAT/CATEGORY/CODE/summary.csv veya
        # fail/REASON/DATE/VIEW/REPEAT/CATEGORY/CODE/summary.csv
        try:
            if parts[0] == 'success' and len(parts) >= 7:
                date, view, repeat, category, code = parts[1], parts[2], parts[3], parts[4], parts[5]
            elif parts[0] == 'fail' and len(parts) >= 8:
                date, view, repeat, category, code = parts[2], parts[3], parts[4], parts[5], parts[6]
            else:
                continue

            # Summary'den hizi oku
            with open(summary_path, 'r', encoding='utf-8') as f:
                for line in f:
                    if line.startswith('Hiz,'):
                        parts_line = line.strip().split(',')
                        if len(parts_line) >= 2:
                            speed = float(parts_line[1])
                            # cm/s ise m/s'e cevir
                            if speed > 1:
                                speed = speed / 100

                            key = f"{view}/{date}/{repeat}/{category}/{code}"
                            speed_cache[key] = speed
                        break
        except Exception as e:
            pass

    print(f"Toplam {len(speed_cache)} hiz degeri yuklendi")

def parse_klasor(klasor_str):
    """Klasor adindan tarih ve view bilgisini cikar"""
    klasor_str = str(klasor_str)

    view = None
    if '(MAK)' in klasor_str or '-MAK' in klasor_str or 'MAK' in klasor_str.upper():
        view = 'MAK'
    elif '(ANG)' in klasor_str or '-ANG' in klasor_str or 'ANG' in klasor_str.upper():
        view = 'ANG'

    date_match = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{2,4})', klasor_str)
    if date_match:
        day, month, year = date_match.groups()
        # S3'deki formata uygun birak - yili degistirme
        date = f"{day}.{month}.{year}"
    else:
        date = None

    return date, view

def get_speed(view, date, repeat, category, code):
    """Cache'den hiz degerini al"""
    key = f"{view}/{date}/{repeat}/{category}/{code}"
    return speed_cache.get(key)

def update_excel():
    build_speed_cache()

    print("\nExcel dosyasi yukleniyor...")
    wb = load_workbook(EXCEL_PATH)

    total_updated = 0
    total_not_found = 0

    for sheet_name in ['MAK', 'ANG']:
        print(f"\n=== {sheet_name} sayfasi isleniyor ===")
        ws = wb[sheet_name]

        headers = [cell.value for cell in ws[1]]
        hiz_col = headers.index('Hiz (m/s)') + 1
        klasor_col = headers.index('Klasor') + 1
        deney_col = headers.index('Deney') + 1 if 'Deney' in headers else None
        kategori_col = headers.index('Kategori') + 1
        kod_col = headers.index('Kod') + 1

        updated = 0
        not_found = 0

        for row in range(2, ws.max_row + 1):
            klasor = ws.cell(row=row, column=klasor_col).value
            deney = ws.cell(row=row, column=deney_col).value if deney_col else 'FIRST'
            kategori = ws.cell(row=row, column=kategori_col).value
            kod = ws.cell(row=row, column=kod_col).value
            old_speed = ws.cell(row=row, column=hiz_col).value

            if not klasor or not kategori or not kod:
                continue

            date, view = parse_klasor(klasor)
            if not date or not view:
                continue

            repeat = str(deney).upper() if deney else 'FIRST'
            new_speed = get_speed(view, date, repeat, kategori, kod)

            if new_speed is not None:
                ws.cell(row=row, column=hiz_col).value = round(new_speed, 4)
                updated += 1
                if updated <= 5:
                    print(f"  {view}/{date}/{repeat}/{kategori}/{kod}: {old_speed} -> {new_speed:.4f}")
            else:
                not_found += 1

        print(f"  Guncellenen: {updated}, Bulunamayan: {not_found}")
        total_updated += updated
        total_not_found += not_found

    # Ortalama sayfalari
    for sheet_name in ['ANG Ortalama', 'MAK Ortalama']:
        print(f"\n=== {sheet_name} sayfasi isleniyor ===")
        ws = wb[sheet_name]

        headers = [cell.value for cell in ws[1]]
        hiz_col = headers.index('Hiz (m/s)') + 1
        klasor_col = headers.index('Klasor') + 1
        kategori_col = headers.index('Kategori') + 1
        kod_col = headers.index('Kod') + 1

        updated = 0
        not_found = 0

        for row in range(2, ws.max_row + 1):
            klasor = ws.cell(row=row, column=klasor_col).value
            kategori = ws.cell(row=row, column=kategori_col).value
            kod = ws.cell(row=row, column=kod_col).value
            old_speed = ws.cell(row=row, column=hiz_col).value

            if not klasor or not kategori or not kod:
                continue

            date, view = parse_klasor(klasor)
            if not date or not view:
                continue

            speeds = []
            for repeat in ['FIRST', 'SECOND', 'THIRD']:
                speed = get_speed(view, date, repeat, kategori, kod)
                if speed is not None:
                    speeds.append(speed)

            if speeds:
                avg_speed = sum(speeds) / len(speeds)
                ws.cell(row=row, column=hiz_col).value = round(avg_speed, 4)
                updated += 1
                if updated <= 3:
                    print(f"  {view}/{date}/{kategori}/{kod}: {old_speed} -> {avg_speed:.4f} (n={len(speeds)})")
            else:
                not_found += 1

        print(f"  Guncellenen: {updated}, Bulunamayan: {not_found}")
        total_updated += updated
        total_not_found += not_found

    print(f"\n=== TOPLAM ===")
    print(f"Guncellenen: {total_updated}")
    print(f"Bulunamayan: {total_not_found}")

    print(f"\nKaydediliyor: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    print("Tamamlandi!")

if __name__ == '__main__':
    update_excel()
